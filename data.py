import pandas as pd
import numpy as np
import torch
import matplotlib.pyplot as plt
import openpyxl
import itertools


DNA_BASES = u"ATGC"

def raw_count(file_name):
    apts = {'A':[1,0,0,0], 'T':[0,1,0,0], 'C':[0,0,1,0], 'G':[0,0,0,1]}
    aptamer_l = []
    f = open(file_name,'r')
    aptamers = f.readlines()
    for aptamer in aptamers:
        apt = []
        aptamer = aptamer.strip('\n').split(',')
        for nuc in aptamer[0]:
            apt.extend(apts[nuc])
        apt.extend( 0 for i in range(200 - len(apt)))
        aptamer_l.append([apt,aptamer[1]])
    print(aptamer_l)

    book_0 = openpyxl.Workbook()
    del book_0["Sheet"]

    sheet = book_0.create_sheet("data")
    row = 1
    for data_0 in aptamer_l:
        sheet.cell(row, 1, data_0[0])
        sheet.cell(row, 2, data_0[1])
        row += 1

    book_0.save('APT_data.xlsx')

def count_all_dna_kmers(sequence_tensor, kmer_k_max):
    kmers , kmers_counts = _kmer_labels(kmer_k_max)
    for i, mer in enumerate(kmers):
        k = sequence_tensor.count(mer)
        kmers_counts[1][i] = k
    return kmers_counts


def _kmer_labels(k_max):  # 获取k-mer，比如 'AT','AG','AC','AA'等
    kmers = []
    for k in range(1, k_max + 1):
        kmers.extend(''.join(s) for s in itertools.product(DNA_BASES, repeat=k))
    kmers_counts_list = kmers.copy()
    kmers_counts_list = [kmers_counts_list]
    kmers_counts_list.append([0 for i in range(len(kmers))])
    return kmers,kmers_counts_list

def _kmer_mean_and_std(kmer_size, sequence_length):
    # assume a binomial distribution
    n = sequence_length + 1 - kmer_size
    p = 1.0 / 4 ** kmer_size
    return n * p, np.sqrt(n * p * (1 - p))

def _all_kmer_mean_and_std(k_max, sequence_length):
  means = []
  stds = []
  for k in range(1, k_max + 1):
    mean, std = _kmer_mean_and_std(k, sequence_length)
    n_kmers = 4 ** k
    means.extend([mean] * n_kmers)
    stds.extend([std] * n_kmers)
  return np.array(means), np.array(stds)


def get_kmer_feature(sequence_tensor, kmer_k_max, sequence_length):
    counts = torch.FloatTensor(count_all_dna_kmers(sequence_tensor, kmer_k_max)[1])
    means, stds = _all_kmer_mean_and_std(kmer_k_max, sequence_length)

    kmer_feature = (( counts - means)/stds)

    return kmer_feature.to(torch.float32)



# data = pd.read_excel('./APT220624.R79-AGS.xlsx')
# for i in data.index:
#     apt = []
#     for nuc in data.loc[i].values[0]:
#         apt.extend(apts[nuc])
#     print(apt)
#     data.loc[i, 'Seq'] = [(apt)]
# print(data.head())

